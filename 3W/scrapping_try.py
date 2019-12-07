from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
import datetime

print(datetime.datetime.now())

file = '/Users/donghyunyoo/Desktop/SCC/Sparta/3W/prac01.xlsx'
work_book = load_workbook(file)
work_sheet = work_book['prac']
# URL을 읽어서 HTML를 받아오고,
headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
data = requests.get('https://movie.naver.com/movie/sdb/rank/rmovie.nhn?sel=pnt&date=20190909',headers=headers)

# HTML을 BeautifulSoup이라는 라이브러리를 활용해 검색하기 용이한 상태로 만듦
soup = BeautifulSoup(data.text, 'html.parser')

# select를 이용해서, tr들을 불러오기
movies = soup.select('#old_content > table > tbody > tr')

# movies (tr들) 의 반복문을 돌리기
rank = 1
work_sheet.cell(row=1,column=1, value="랭킹")
work_sheet.cell(row=1,column=2, value="타이틀")
work_sheet.cell(row=1,column=3, value="별점")
work_sheet.delete_cols(4)
row = 2

for movie in movies:
    # movie 안에 a 가 있으면,
    a_tag = movie.select_one('td.title > div > a')
    if a_tag is not None:
        title = a_tag.text
        star = movie.select_one('td.point').text
        work_sheet.cell(row=row,column=1, value=rank)
        work_sheet.cell(row=row,column=2, value=title)
        work_sheet.cell(row=row, column=3, value=star)
        rank += 1
        row+=1

work_book.save(file)
print(datetime.datetime.now())
